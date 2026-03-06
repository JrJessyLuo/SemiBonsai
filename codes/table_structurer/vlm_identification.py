import argparse
import glob
import json
import os
import re
import time
import traceback
from pathlib import Path
from typing import Dict
import openpyxl
import tqdm
import yaml
from loguru import logger
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from utils.extract_excel import process_table_vlm as process_table_vlm_multiple
from utils.api_utils import vlm_generate_fewshot_setup
from utils.constants import DELIMITER
from utils.prompt_template import identify_subtable_prompt, identify_subtable_prompt_multiple
from utils.sheet_utils import *
from utils.attr_extraction import sheet_to_image


def save_sheet_raw_format(tab_sheet: Worksheet, out_xlsx_path: str | Path) -> None:
    """
    Copy ONLY:
      - cell values
      - merged cell ranges
    This is the most robust way to preserve the main content layout.
    """
    out_xlsx_path = str(out_xlsx_path)

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = tab_sheet.title

    # 1) copy values (skip MergedCell objects)
    for row in tab_sheet.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

    # 2) copy merged ranges (must be done after values)
    for merged_range in tab_sheet.merged_cells.ranges:
        ws_new.merge_cells(str(merged_range))

    wb_new.save(out_xlsx_path)


def flatten_sheet_dict(nested_dict, default_prefix="subtab") -> Dict[str, Worksheet]:
    """
    Flatten a nested dict into {leaf_key: Worksheet}.
    - Keeps only Worksheet leaves
    - Auto-names empty keys: subtab_1, subtab_2, ...
    - Resolves collisions by suffixing _1, _2, ...
    """
    flat: Dict[str, Worksheet] = {}
    unnamed_counter = 1

    def _walk(d):
        nonlocal unnamed_counter
        for k, v in d.items():
            if isinstance(v, dict):
                _walk(v)
                continue

            if isinstance(v, Worksheet):
                name = (k or "").strip()
                if not name:
                    name = f"{default_prefix}_{unnamed_counter}"
                    unnamed_counter += 1

                orig_name = str(name)
                name = orig_name
                idx = 1
                while name in flat:
                    name = f"{orig_name}_{idx}"
                    idx += 1

                flat[name] = v

    if isinstance(nested_dict, dict):
        _walk(nested_dict)

    return flat


def preprocess_cell(value):
    return str(value)


def preprocess_sheet(sheet: Worksheet) -> Worksheet:
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.value = preprocess_cell(cell.value)
    return sheet


def match_minimal_table_structure(sheet: Worksheet):
    """
    Match minimal structure (one or two merged blocks).
    Returns dict if matched; otherwise None.
    """
    nrows = sheet.max_row
    ncols = sheet.max_column

    first_cell = sheet.cell(row=1, column=1)
    _, _, x2, y2 = get_merge_cell_size(sheet, first_cell.coordinate)

    if x2 == nrows and y2 == ncols:
        return {get_merge_cell_value(sheet, first_cell.coordinate): None}

    if x2 == nrows:  # two horizontal blocks
        second_cell = sheet.cell(row=1, column=y2 + 1)
        _, _, xx2, yy2 = get_merge_cell_size(sheet, second_cell.coordinate)
        if yy2 == ncols:
            return {
                get_merge_cell_value(sheet, first_cell.coordinate): get_merge_cell_value(
                    sheet, second_cell.coordinate
                )
            }

    if y2 == ncols:  # two vertical blocks
        second_cell = sheet.cell(row=x2 + 1, column=1)
        _, _, xx2, yy2 = get_merge_cell_size(sheet, second_cell.coordinate)
        if xx2 == nrows:
            return {
                get_merge_cell_value(sheet, first_cell.coordinate): get_merge_cell_value(
                    sheet, second_cell.coordinate
                )
            }

    return None


def get_nrow_cells(sheet: Worksheet) -> int:
    n = 0
    nrows, ncols = sheet.max_row, sheet.max_column

    for row in range(1, nrows + 1):
        col = 1
        cnt = 0
        while col <= ncols:
            cnt += 1
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            col += (y2 - y1 + 1)
        n = max(n, cnt)

    return n


def get_ncol_cells(sheet: Worksheet) -> int:
    n = 0
    nrows, ncols = sheet.max_row, sheet.max_column

    for col in range(1, ncols + 1):
        row = 1
        cnt = 0
        while row <= nrows:
            cnt += 1
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            row += (x2 - x1 + 1)
        n = max(n, cnt)

    return n


def process_sheet_vlm_in(
    sheet: Worksheet,
    args,
    single_table=True,
):
    if sheet is None:
        logger.info("Sheet is None, return.")
        return None, 0, 0

    # Step 1: delete empty rows/cols
    delete_empty_columns(sheet)
    delete_empty_rows(sheet)

    nrows, ncols = sheet.max_row, sheet.max_column
    nrow_cells = get_nrow_cells(sheet)
    ncol_cells = get_ncol_cells(sheet)

    logger.info(f"{DELIMITER} Enter process_sheet_vlm_in() {DELIMITER}")
    logger.info(
        f"nrows={nrows} ncols={ncols}; "
        f"max_cells_per_row={nrow_cells} max_cells_per_col={ncol_cells}"
    )

    # Early return: minimal structure
    res = match_minimal_table_structure(sheet)
    logger.info(f"{DELIMITER} Early return check (minimal structure) {DELIMITER}")
    if res is not None:
        logger.info(f"Matched minimal structure: {res}")
        return res, 0, 0

    # Convert to image + VLM
    image_file = sheet_to_image(sheet)
    logger.info(f"{DELIMITER} Sheet image path {DELIMITER}\n{image_file}")

    if single_table:
        output = vlm_generate_fewshot_setup(
            identify_subtable_prompt, image_file, json_format=True, single_tab=True
        )
    else:
        output = vlm_generate_fewshot_setup(
            identify_subtable_prompt_multiple, image_file, json_format=True, single_tab=False
        )


    identified_subtables = output["text"]
    return identified_subtables, output.get("input_tokens", 0), output.get("output_tokens", 0)


def process_table_vlm_in(file, args, single_table=True):
    wb = openpyxl.load_workbook(file, data_only=True)
    sheet = preprocess_sheet(wb.active)
    res, input_toks, output_toks = process_sheet_vlm_in(
        sheet, args, single_table=single_table
    )
    return res, input_toks, output_toks


def process_single_table(table_id: str, fp: str, single_table_flag: bool):
    global all_cnt, multiple_subtable_cnt, logs_path, meta_path, multitab_writer, process_cost_writer

    log_file = os.path.join(logs_path, f"{table_id}.log")
    meta_file = os.path.join(meta_path, f"{table_id}.json")

    try:
        log_file_handler = logger.add(log_file)

        start_time = time.time()
        table_infos, input_toks, output_toks = process_table_vlm_in(
            fp,
            args,
            single_table=single_table_flag,
        )
        cost_time = time.time() - start_time

        logger.info(f"Cost time: {cost_time:.3f}s")
        logger.remove(log_file_handler)

        all_cnt += 1

        # table_infos might be a dict or other JSON
        if isinstance(table_infos, dict):
            if str(table_infos.get("single_subtable", "")).lower() == "no":
                multiple_subtable_cnt += 1
                if multitab_writer:
                    multitab_writer.write(table_id + "\n")
                    multitab_writer.flush()

        with open(meta_file, "w", encoding="utf-8") as f_meta:
            json.dump(table_infos, f_meta, ensure_ascii=False, indent=2)

        cost_rec = {
            "table_id": table_id,
            "time_cost": cost_time,
            "token_cost": [input_toks, output_toks],
        }
        process_cost_writer.write(json.dumps(cost_rec, ensure_ascii=False) + "\n")
        process_cost_writer.flush()

    except Exception as e:
        err = "".join(traceback.format_exception(type(e), e, e.__traceback__))
        with open(log_file, "w", encoding="utf-8") as f_err:
            f_err.write(f"[ERROR] {table_id}\n{err}\n")


def parse_option():
    parser = argparse.ArgumentParser("command line arguments for generation.")
    parser.add_argument("--dataset", type=str, required=True, help="dataset name")
    parser.add_argument("--multiple", action="store_true", help="use multiple-subtable splitting mode")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_option()

    with open("../config.yaml", "r") as f:
        cfg = yaml.safe_load(f)

    base_folder = cfg["base_folder"]
    out_folder = cfg["out_folder"]
    ds_cfg = cfg["datasets"][args.dataset]

    # single-table mode selection
    single_tab_mode = (("hitab_num" in args.dataset) or ("multihiertt_num" in args.dataset)) and (not args.multiple)

    basic_fpath = os.path.join(base_folder, args.dataset, ds_cfg["table_dir"])
    print(basic_fpath)

    logs_path = os.path.join(out_folder, args.dataset, "our/table_processed/logs")
    meta_path = os.path.join(out_folder, args.dataset, "our/table_processed/meta_infos")
    os.makedirs(logs_path, exist_ok=True)
    os.makedirs(meta_path, exist_ok=True)

    multitab_path = os.path.join(out_folder, args.dataset, "our/table_processed/multitab_process.txt")

    # JSONL file for time & token costs
    cost_path = os.path.join(out_folder, args.dataset, "our/table_processed/offline_process.jsonl")
    process_cost_writer = open(cost_path, "a", encoding="utf-8")

    all_cnt = 0
    multiple_subtable_cnt = 0

    processed_ids = []
    multitab_mapping_write = None
    multitab_writer = None

    if args.multiple:
        # only process tables listed in multitab_process.txt
        with open(multitab_path, "r", encoding="utf-8") as f:
            table_ids = [line.strip() for line in f if line.strip()]
        xlsx_paths = [os.path.join(basic_fpath, f"{tid}.xlsx") for tid in table_ids]

        multitab_mapping_fpath = os.path.join(out_folder, args.dataset, "our/table_processed/multitab_mapping.jsonl")
        if os.path.exists(multitab_mapping_fpath):
            with open(multitab_mapping_fpath, "r", encoding="utf-8") as f:
                processed_ids = [json.loads(line)["raw_table_id"] for line in f if line.strip()]

        multitab_mapping_write = open(multitab_mapping_fpath, "a", encoding="utf-8")
        multitab_writer = open(multitab_path, "a", encoding="utf-8")

    else:
        # process all tables; record multi-subtable ones into multitab_process.txt
        xlsx_paths = sorted(glob.glob(os.path.join(basic_fpath, "*.xlsx")))
        multitab_writer = open(multitab_path, "a", encoding="utf-8")

    for fp in tqdm.tqdm(xlsx_paths, total=len(xlsx_paths), desc="Processing tables"):
        raw_table_id = os.path.splitext(os.path.basename(fp))[0]

        if args.multiple and raw_table_id in processed_ids:
            continue

        if args.multiple:
            print(f"Begin to process raw table: {raw_table_id}")
            raw_output = process_table_vlm_multiple(fp, col_span=True)

            output = flatten_sheet_dict(raw_output)

            sheet_mapping_dict = {raw_table_id: []}
            sheet_fpath_dict = {}

            for tab_key, tab_sheet in output.items():
                if tab_sheet is None:
                    continue

                safe_key = re.sub(r"[\\/]+", "_", str(tab_key).strip())
                updated_tab_key = f"subtab_{raw_table_id}_{safe_key}"
                if len(updated_tab_key) >= 80:
                    updated_tab_key = updated_tab_key[:80]

                sheet_mapping_dict[raw_table_id].append(updated_tab_key)

                out_xlsx_path = os.path.join(basic_fpath, f"{updated_tab_key}.xlsx")
                save_sheet_raw_format(tab_sheet, out_xlsx_path)
                sheet_fpath_dict[updated_tab_key] = out_xlsx_path

            if sheet_mapping_dict[raw_table_id]:
                mapping_record = {"raw_table_id": raw_table_id, "subtab_ids": sheet_mapping_dict[raw_table_id]}
                multitab_mapping_write.write(json.dumps(mapping_record, ensure_ascii=False) + "\n")
                multitab_mapping_write.flush()

            for subtab_id, sub_fp in sheet_fpath_dict.items():
                process_single_table(
                    table_id=subtab_id,
                    fp=sub_fp,
                    single_table_flag=True,
                )

        else:
            # skip already split subtables
            if raw_table_id.startswith("subtab_"):
                continue

            process_single_table(
                table_id=raw_table_id,
                fp=fp,
                single_table_flag=single_tab_mode,
            )

    if multitab_writer:
        multitab_writer.close()
    if multitab_mapping_write:
        multitab_mapping_write.close()
    process_cost_writer.close()

    pct = (multiple_subtable_cnt / all_cnt * 100) if all_cnt else 0.0
    print(f"Complex multiple-subtable tables: {multiple_subtable_cnt}/{all_cnt} ({pct:.1f}%)")