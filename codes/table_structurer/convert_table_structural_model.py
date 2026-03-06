# -*- coding: utf-8 -*-
import argparse
import hashlib
import json
import os
import re
from typing import Any, Dict, List, Optional, Tuple

import networkx as nx
import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


# ==============================
# Utils: trimming empty margins
# ==============================
def _cell_is_empty(val: Any) -> bool:
    if val is None:
        return True
    if isinstance(val, str) and val.strip() == "":
        return True
    return False


def clean_worksheet(ws: Worksheet, max_rows: Optional[int] = None) -> Worksheet:
    max_row_full = ws.max_row
    max_col_full = ws.max_column

    first_row = None
    last_row = None
    for r in range(1, max_row_full + 1):
        if any(not _cell_is_empty(ws.cell(row=r, column=c).value) for c in range(1, max_col_full + 1)):
            first_row = r
            break
    for r in range(max_row_full, 0, -1):
        if any(not _cell_is_empty(ws.cell(row=r, column=c).value) for c in range(1, max_col_full + 1)):
            last_row = r
            break

    first_col = None
    last_col = None
    for c in range(1, max_col_full + 1):
        if any(not _cell_is_empty(ws.cell(row=r, column=c).value) for r in range(1, max_row_full + 1)):
            first_col = c
            break
    for c in range(max_col_full, 0, -1):
        if any(not _cell_is_empty(ws.cell(row=r, column=c).value) for r in range(1, max_row_full + 1)):
            last_col = c
            break

    if first_row is None or first_col is None:
        return ws

    if max_rows is not None:
        visible_rows = last_row - first_row + 1
        if visible_rows > max_rows:
            last_row = first_row + max_rows - 1

    for r in range(max_row_full, last_row, -1):
        ws.delete_rows(r)
    for r in range(first_row - 1, 0, -1):
        ws.delete_rows(r)

    for c in range(max_col_full, last_col, -1):
        ws.delete_cols(c)
    for c in range(first_col - 1, 0, -1):
        ws.delete_cols(c)

    return ws


# ==============================
# Utils: meta -> leaf paths
# ==============================
def enumerate_leaf_paths(groups) -> List[List[str]]:
    """
    Convert header group tree into list of root-to-leaf label paths.
    Supports:
      - dict node: {"group":..., "children":[...]}
      - scalar leaf: "xxx"
      - dict-leaf: {"group":"xxx","children":[]}
    """
    paths: List[List[str]] = []

    def dfs(node, prefix):
        if isinstance(node, dict):
            g = node.get("group", None)
            new_prefix = prefix + ([str(g)] if g not in (None, "") else [])
            kids = node.get("children") or []
            if not kids:
                if g not in (None, ""):
                    paths.append(new_prefix)
                return
            for ch in kids:
                dfs(ch, new_prefix)
        elif isinstance(node, (list, tuple)):
            for ch in node:
                dfs(ch, prefix)
        else:
            leaf = str(node).strip()
            if leaf:
                paths.append(prefix + [leaf])

    dfs(groups, [])
    return paths


def make_path_id(subtab_id: str, axis: str, path: List[str]) -> str:
    h = hashlib.md5(("||".join(path)).encode("utf-8")).hexdigest()[:10]
    return f"{subtab_id}::{axis}::{h}"


# ==============================
# Multi-subtable mapping helpers
# ==============================
def load_multitab_mapping_jsonl(path: str) -> Dict[str, List[str]]:
    """
    multitab_mapping.jsonl line format:
      {"raw_table_id": "...", "subtab_ids": [...]}
    Returns raw_table_id -> subtab_ids
    """
    m: Dict[str, List[str]] = {}
    if not os.path.exists(path):
        return m
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rec = json.loads(line)
            raw = rec.get("raw_table_id")
            subs = rec.get("subtab_ids", [])
            if raw:
                m[raw] = subs
    return m


def infer_raw_table_id(table_id: str) -> str:
    # naming: subtab_{raw}_...
    if table_id.startswith("subtab_"):
        parts = table_id.split("_", 2)
        if len(parts) >= 2:
            return parts[1]
    return table_id


def infer_subtable_title_from_id(subtable_id: str) -> Optional[str]:
    """
    subtab_religion-table07_2012 -> "2012"
    If not subtab_*, return None.
    """
    if not subtable_id.startswith("subtab_"):
        return None
    parts = subtable_id.split("_")
    return parts[-1] if len(parts) >= 2 else None


def build_raw_to_group(all_ids: List[str], raw2subs: Dict[str, List[str]]) -> Dict[str, List[str]]:
    """
    Build raw_table_id -> [subtable_ids].
    Rules:
      - If raw has any subtab_* children, do NOT include raw itself as a subtable.
      - If raw has no children, treat raw itself as a single subtable.
      - Add safety: include any subtab_* present in meta_dir even if not in mapping.
    """
    raw_to_group = {raw: list(subs) for raw, subs in raw2subs.items()}

    # add any subtab_* not listed (safety)
    for tid in all_ids:
        if tid.startswith("subtab_"):
            raw = infer_raw_table_id(tid)
            raw_to_group.setdefault(raw, [])
            if tid not in raw_to_group[raw]:
                raw_to_group[raw].append(tid)

    # add single tables that have NO subtab children
    for tid in all_ids:
        if tid.startswith("subtab_"):
            continue
        if tid not in raw_to_group or not raw_to_group[tid]:
            raw_to_group.setdefault(tid, [])
            if tid not in raw_to_group[tid]:
                raw_to_group[tid].append(tid)

    # IMPORTANT: if a raw table has any subtab_* children, drop raw itself if included
    for raw, subs in list(raw_to_group.items()):
        has_child = any(s.startswith("subtab_") for s in subs)
        if has_child:
            raw_to_group[raw] = [s for s in subs if s != raw]

    # dedup stable
    for raw in raw_to_group:
        raw_to_group[raw] = list(dict.fromkeys(raw_to_group[raw]))
    return raw_to_group


# ==============================
# Graph helpers (for value indexing)
# ==============================
def _canon(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    t = re.sub(r"\s+", "", t)
    return t


def build_merged_maps(ws: Worksheet):
    merged_top = {}
    merged_blocks = []
    for mr in ws.merged_cells.ranges:
        try:
            r0, c0, r1, c1 = mr.min_row, mr.min_col, mr.max_row, mr.max_col
        except Exception:
            try:
                r0, c0, r1, c1 = range_boundaries(str(mr))
            except Exception:
                continue
        merged_blocks.append((r0, c0, r1, c1))
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                merged_top[(r, c)] = (r0, c0)
    return merged_top, merged_blocks


def merged_value(ws: Worksheet, r: int, c: int, merged_top: Dict[Tuple[int, int], Tuple[int, int]]):
    r0, c0 = merged_top.get((r, c), (r, c))
    return ws.cell(r0, c0).value


def _add_header_leaf(
    G: nx.Graph,
    kind: str,
    node_id: str,
    label: str,
    coord: Optional[Tuple[int, int]] = None,
    path: Optional[List[str]] = None,
    is_group: bool = False
):
    attrs = {"type": kind, "label": label, "is_group": is_group}
    if coord:
        attrs["coord"] = f"R{coord[0]}C{coord[1]}"
    if path is not None:
        attrs["path"] = list(path)
    G.add_node(node_id, **attrs)


def _add_value(G: nx.Graph, node_id: str, val: Any, r: int, c: int):
    G.add_node(node_id, type="value", label=str(val), coord=f"R{r}C{c}")


def build_graph_from_metadata(
    xlsx_path: str,
    structure: Dict[str, Any],
    ignore_sparse_rows: bool = False,
    sparse_row_min_vals: int = 2,
    restrict_to_below_header: bool = True,
    collapse_headers: bool = True,
    path_sep: str = ", ",
    use_distinct_labels: bool = True,
) -> Optional[nx.Graph]:
    """
    Minimal graph builder:
      - Column header leaf nodes store full path (list[str])
      - Row header leaf nodes store full path (list[str]) via meta lookup
      - Value nodes connect to (row_leaf, col_leaf)
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = clean_worksheet(wb.active, max_rows=200)
    merged_top, _ = build_merged_maps(ws)

    G = nx.Graph()

    # ----- meta paths (for row header) -----
    has_row_header = str(structure.get("has_row_header", "no")).lower() == "yes"
    row_groups = structure.get("row_header_groups", []) if has_row_header else []
    row_leaf_paths = enumerate_leaf_paths(row_groups) if has_row_header else []
    leaflabel2rowpaths: Dict[str, List[List[str]]] = {}
    for p in row_leaf_paths:
        if p:
            leaflabel2rowpaths.setdefault(p[-1], []).append(p)

    # ----- col groups -----
    col_groups = structure.get("column_header_groups", [])
    col_leaf_paths = enumerate_leaf_paths(col_groups)
    leaflabel2colpaths: Dict[str, List[List[str]]] = {}
    for p in col_leaf_paths:
        if p:
            leaflabel2colpaths.setdefault(p[-1], []).append(p)

    # ----- locate leaf header row (heuristic) -----
    leaf_labels_norm = set(_canon(p[-1]) for p in col_leaf_paths if p)
    hdr_leaf_row = None
    best_cnt = -1
    for r in range(1, ws.max_row + 1):
        cnt = 0
        seen = set()
        for c in range(1, ws.max_column + 1):
            r0, c0 = merged_top.get((r, c), (r, c))
            if (r0, c0) in seen:
                continue
            seen.add((r0, c0))
            v = ws.cell(r0, c0).value
            if _canon(v) in leaf_labels_norm:
                cnt += 1
        if cnt > best_cnt:
            best_cnt = cnt
            hdr_leaf_row = r
    if hdr_leaf_row is None:
        return None

    header_top = 1

    # ----- Build column leaf nodes per column index -----
    path_node_by_prefix: Dict[Tuple[str, ...], str] = {}
    col_path_nodes_by_col: Dict[int, List[Tuple[str, bool]]] = {}

    def ensure_col_path_nodes_for_col(col_path: List[str], col_idx: int):
        nodes_chain: List[Tuple[str, bool]] = []
        parent_path = col_path[:-1]
        leaf = col_path[-1]

        # ancestors
        for level in range(1, len(parent_path) + 1):
            prefix = tuple(parent_path[:level])
            if prefix not in path_node_by_prefix:
                label = path_sep.join(list(prefix)) if collapse_headers else prefix[-1]
                level_row = header_top + (level - 1)
                nid = f"col::anc::{path_sep.join(prefix)}"
                _add_header_leaf(G, "column_header", nid, label=label, coord=(level_row, col_idx), path=list(prefix))
                path_node_by_prefix[prefix] = nid
            nodes_chain.append((path_node_by_prefix[prefix], False))

        # leaf
        leaf_path = tuple(col_path)
        if leaf_path not in path_node_by_prefix:
            label = path_sep.join(list(leaf_path)) if collapse_headers and parent_path else leaf
            nid = f"col::leaf::{label}"
            _add_header_leaf(G, "column_header", nid, label=label, coord=(hdr_leaf_row, col_idx), path=list(leaf_path))
            path_node_by_prefix[leaf_path] = nid
        nodes_chain.append((path_node_by_prefix[leaf_path], True))
        return nodes_chain

    for c in range(1, ws.max_column + 1):
        v = merged_value(ws, hdr_leaf_row, c, merged_top)
        leaf = str(v).strip() if v not in (None, "") else ""
        if not leaf:
            continue
        cands = leaflabel2colpaths.get(leaf, [])
        if not cands:
            continue
        col_path = cands[0]
        chain = ensure_col_path_nodes_for_col(col_path, c)
        col_path_nodes_by_col[c] = chain

    if not col_path_nodes_by_col:
        return None

    # ----- Infer row header column -----
    row_header_col = None
    if has_row_header and row_leaf_paths:
        counts: Dict[int, int] = {}
        row_leaf_set = set(_canon(p[-1]) for p in row_leaf_paths if p)
        for r in range(hdr_leaf_row + 1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = merged_value(ws, r, c, merged_top)
                if _canon(v) in row_leaf_set:
                    counts[c] = counts.get(c, 0) + 1
        if counts:
            row_header_col = min([c for c, k in counts.items() if k == max(counts.values())])
    if row_header_col is None:
        row_header_col = 1

    def _row_is_sparse(r: int) -> bool:
        if not ignore_sparse_rows:
            return False
        cnt = 0
        seen_tops = set()
        for c in range(1, ws.max_column + 1):
            if c == row_header_col:
                continue
            r0, c0 = merged_top.get((r, c), (r, c))
            if (r0, c0) in seen_tops:
                continue
            seen_tops.add((r0, c0))
            v = ws.cell(r0, c0).value
            if v not in (None, ""):
                cnt += 1
                if cnt >= sparse_row_min_vals:
                    return False
        return True

    # ----- Build row header leaf nodes per row -----
    row_leaf_for_row: Dict[int, str] = {}
    current_band_label = None
    for r in range(hdr_leaf_row + 1, ws.max_row + 1):
        v = merged_value(ws, r, row_header_col, merged_top)

        if (not has_row_header) and _row_is_sparse(r):
            current_band_label = v

        if v not in (None, ""):
            lbl = str(v).strip()

            band = str(current_band_label).strip() if current_band_label not in (None, "") else ""
            if band and band != lbl:
                label = f"{band}, {lbl}"
                row_path = [band, lbl]
            else:
                label = lbl
                cands = leaflabel2rowpaths.get(lbl, [])
                row_path = cands[0] if cands else [lbl]

            nid = f"row::leaf::{lbl}::R{r}C{row_header_col}"
            base, k = nid, 2
            while nid in G:
                nid = f"{base}__{k}"
                k += 1

            _add_header_leaf(G, "row_header", nid, label=label, coord=(r, row_header_col), path=row_path)
            row_leaf_for_row[r] = nid
        else:
            syn = f"row::syn::R{r}"
            if syn not in G:
                band = str(current_band_label).strip() if current_band_label not in (None, "") else ""
                syn_label = f"Row {r}"
                syn_path = [band, syn_label] if band else [syn_label]
                _add_header_leaf(G, "row_header", syn, label=syn_label, coord=(r, 0), path=syn_path)
            row_leaf_for_row[r] = syn

    # ----- Emit values -----
    def connect_value(vid: str, col_chain, row_leaf):
        for nid, is_leaf in col_chain:
            if is_leaf or not collapse_headers:
                G.add_edge(vid, nid)
        if row_leaf:
            G.add_edge(vid, row_leaf)

    row_iter = range(hdr_leaf_row + 1, ws.max_row + 1) if restrict_to_below_header else range(1, ws.max_row + 1)
    for r in row_iter:
        if _row_is_sparse(r):
            continue
        rleaf = row_leaf_for_row.get(r)
        if not rleaf:
            continue

        for c, col_chain in col_path_nodes_by_col.items():
            if c == row_header_col:
                continue
            v = merged_value(ws, r, c, merged_top)
            if v in (None, ""):
                continue
            vid = f"val::R{r}C{c}"
            _add_value(G, vid, v, r, c)
            connect_value(vid, col_chain, rleaf)

    return G


# ==============================
# Layered Tree (nodes+edges) builder
# ==============================
def build_header_tree_nodes_edges(subtable_id: str, groups, axis: str):
    """
    axis: "C" for column, "R" for row
    Returns: (nodes, edges)
    Node schema:
      - id: deterministic hash id from path
      - type: "colhdr" | "rowhdr"
      - label
      - is_leaf
      - path (list[str])
      - path_id (only for leaf)
    """
    nodes: List[Dict[str, Any]] = []
    edges: List[List[str]] = []

    def node_id(path: List[str]) -> str:
        prefix = "c" if axis == "C" else "r"
        h = hashlib.md5(("||".join(path)).encode("utf-8")).hexdigest()[:10]
        return f"{prefix}::{subtable_id}::{h}"

    path2id: Dict[Tuple[str, ...], str] = {}

    def ensure_node(path: List[str], label: str, is_leaf: bool) -> str:
        key = tuple(path)
        if key in path2id:
            return path2id[key]
        nid = node_id(path)
        nd = {
            "id": nid,
            "type": ("colhdr" if axis == "C" else "rowhdr"),
            "label": label,
            "is_leaf": is_leaf,
            "path": path,
        }
        if is_leaf:
            nd["path_id"] = make_path_id(subtable_id, axis, path)
        nodes.append(nd)
        path2id[key] = nid
        return nid

    def dfs(node, prefix: List[str], parent_id: Optional[str]):
        if isinstance(node, dict):
            g = node.get("group", None)
            kids = node.get("children") or []

            if g in (None, ""):
                for ch in kids:
                    dfs(ch, prefix, parent_id)
                return

            cur_path = prefix + [str(g).strip()]
            if not kids:
                cur_id = ensure_node(cur_path, label=str(g).strip(), is_leaf=True)
                if parent_id:
                    edges.append([parent_id, cur_id])
                return

            cur_id = ensure_node(cur_path, label=str(g).strip(), is_leaf=False)
            if parent_id:
                edges.append([parent_id, cur_id])
            for ch in kids:
                dfs(ch, cur_path, cur_id)

        elif isinstance(node, (list, tuple)):
            for ch in node:
                dfs(ch, prefix, parent_id)
        else:
            leaf = str(node).strip()
            if not leaf:
                return
            leaf_path = prefix + [leaf]
            leaf_id = ensure_node(leaf_path, label=leaf, is_leaf=True)
            if parent_id:
                edges.append([parent_id, leaf_id])

    for g in groups or []:
        dfs(g, [], None)

    return nodes, edges


def write_layered_tree_for_raw(raw_table_id: str, subtable_ids: List[str], meta_dir: str, out_dir: str):
    """
    Write ONE layered_tree.json per raw table, with explicit nodes+edges AND per-node parent_id.
    """
    os.makedirs(out_dir, exist_ok=True)

    nodes: List[Dict[str, Any]] = [{"id": "root", "type": "root"}]
    edges: List[List[str]] = []

    # We'll maintain a set of node ids to avoid duplicates
    node_id_set = {"root"}

    def add_node(nd: Dict[str, Any]):
        if nd["id"] in node_id_set:
            return
        nodes.append(nd)
        node_id_set.add(nd["id"])

    def add_edge(p: str, c: str):
        edges.append([p, c])

    for sid in subtable_ids:
        meta_path = os.path.join(meta_dir, f"{sid}.json")
        if not os.path.exists(meta_path):
            continue
        meta = json.load(open(meta_path, "r", encoding="utf-8"))

        title = infer_subtable_title_from_id(sid)

        st_id = f"st::{sid}"
        add_node({"id": st_id, "type": "subtable", "subtable_id": sid, "title": title})
        add_edge("root", st_id)

        # Column header tree
        col_groups = meta.get("column_header_groups", [])
        col_nodes, col_edges = build_header_tree_nodes_edges(sid, col_groups, axis="C")

        # connect subtable -> column roots (indeg=0 within col_edges)
        col_indeg: Dict[str, int] = {}
        for p, c in col_edges:
            col_indeg[c] = col_indeg.get(c, 0) + 1
            col_indeg.setdefault(p, col_indeg.get(p, 0))
        col_root_ids = [n["id"] for n in col_nodes if col_indeg.get(n["id"], 0) == 0]

        for nd in col_nodes:
            add_node(nd)
        for p, c in col_edges:
            add_edge(p, c)
        for rid in col_root_ids:
            add_edge(st_id, rid)

        # Row header tree
        has_row = str(meta.get("has_row_header", "no")).lower() == "yes"
        row_groups = meta.get("row_header_groups", []) if has_row else []
        row_nodes, row_edges = build_header_tree_nodes_edges(sid, row_groups, axis="R")

        row_indeg: Dict[str, int] = {}
        for p, c in row_edges:
            row_indeg[c] = row_indeg.get(c, 0) + 1
            row_indeg.setdefault(p, row_indeg.get(p, 0))
        row_root_ids = [n["id"] for n in row_nodes if row_indeg.get(n["id"], 0) == 0]

        for nd in row_nodes:
            add_node(nd)
        for p, c in row_edges:
            add_edge(p, c)
        for rid in row_root_ids:
            add_edge(st_id, rid)

    # ---- Build parent_id and children_ids from edges ----
    parent_of: Dict[str, str] = {}
    children_of: Dict[str, List[str]] = {}

    for p, c in edges:
        # record parent
        if c in parent_of and parent_of[c] != p:
            # This "should not happen" in a tree; keep the first and ignore later.
            # If you later want DAG support, change parent_id -> parent_ids (list).
            pass
        else:
            parent_of[c] = p

        children_of.setdefault(p, []).append(c)

    # attach parent_id and children_ids to each node
    for nd in nodes:
        nid = nd["id"]
        nd["parent_id"] = parent_of.get(nid)  # root will be None
        nd["children_ids"] = children_of.get(nid, [])

    layered_tree = {
        "raw_table_id": raw_table_id,
        "nodes": nodes,
        "edges": edges,
    }
    with open(os.path.join(out_dir, "layered_tree.json"), "w", encoding="utf-8") as f:
        json.dump(layered_tree, f, ensure_ascii=False, indent=2)


# ==============================
# Raw-level Value Index Structure (ONE per raw table)
# ==============================
def build_value_index_for_raw(
    raw_table_id: str,
    subtable_ids: List[str],
    meta_dir: str,
    xlsx_dir: str,
    out_dir: str,
    use_distinct_labels: bool = True,
):
    """
    Write ONE value_table.csv + mapping_table.csv for raw_table_id (even with multiple subtables).
    value_id is int, unique within raw table.

    mapping_table uses col_path_id/row_path_id that encode subtable_id, so paths from different subtables don't collide.
    """
    os.makedirs(out_dir, exist_ok=True)

    value2id: Dict[str, int] = {}
    value_rows: List[Dict[str, Any]] = []
    mapping_rows: List[Dict[str, Any]] = []
    next_vid = 0

    for sid in subtable_ids:
        meta_path = os.path.join(meta_dir, f"{sid}.json")
        xlsx_path = os.path.join(xlsx_dir, f"{sid}.xlsx")
        if not (os.path.exists(meta_path) and os.path.exists(xlsx_path)):
            continue

        meta = json.load(open(meta_path, "r", encoding="utf-8"))

        G = build_graph_from_metadata(
            xlsx_path,
            meta,
            ignore_sparse_rows=False,
            sparse_row_min_vals=2,
            restrict_to_below_header=True,
            collapse_headers=True,
            path_sep=", ",
            use_distinct_labels=use_distinct_labels,
        )
        if G is None:
            # fallback: no row header
            meta2 = dict(meta)
            meta2["has_row_header"] = "no"
            meta2["row_header_groups"] = []
            G = build_graph_from_metadata(
                xlsx_path,
                meta2,
                ignore_sparse_rows=False,
                sparse_row_min_vals=2,
                restrict_to_below_header=True,
                collapse_headers=True,
                path_sep=", ",
                use_distinct_labels=use_distinct_labels,
            )
        if G is None:
            continue

        has_row_header = str(meta.get("has_row_header", "no")).lower() == "yes"
        col_paths = enumerate_leaf_paths(meta.get("column_header_groups", []))
        row_paths = enumerate_leaf_paths(meta.get("row_header_groups", [])) if has_row_header else []

        col_path_id_map = {tuple(p): make_path_id(sid, "C", p) for p in col_paths}
        row_path_id_map = {tuple(p): make_path_id(sid, "R", p) for p in row_paths}

        for n, d in G.nodes(data=True):
            if d.get("type") != "value":
                continue

            col_leaf = None
            row_leaf = None
            for nb in G.neighbors(n):
                t = G.nodes[nb].get("type")
                if t == "column_header" and not G.nodes[nb].get("is_group", False):
                    col_leaf = nb
                elif t == "row_header" and not G.nodes[nb].get("is_group", False):
                    row_leaf = nb
            if col_leaf is None or row_leaf is None:
                continue

            col_path = G.nodes[col_leaf].get("path") or []
            row_path = G.nodes[row_leaf].get("path") or []

            col_pid = col_path_id_map.get(tuple(col_path)) or make_path_id(sid, "C", col_path)
            row_pid = row_path_id_map.get(tuple(row_path)) or make_path_id(sid, "R", row_path)

            val = str(d.get("label", ""))

            if val not in value2id:
                value2id[val] = next_vid
                value_rows.append({"value_id": next_vid, "value": val})
                next_vid += 1

            mapping_rows.append({"col_path_id": col_pid, "row_path_id": row_pid, "value_id": value2id[val]})

    value_df = pd.DataFrame(value_rows)
    mapping_df = pd.DataFrame(mapping_rows).drop_duplicates(subset=["col_path_id", "row_path_id"])

    value_df.to_csv(os.path.join(out_dir, "value_table.csv"), index=False)
    mapping_df.to_csv(os.path.join(out_dir, "mapping_table.csv"), index=False)


# ==============================
# Batch driver
# ==============================
def ensure_dir(p: str):
    os.makedirs(p, exist_ok=True)


def list_all_table_ids(meta_dir: str) -> List[str]:
    ids = []
    for name in os.listdir(meta_dir):
        if name.endswith(".json"):
            ids.append(os.path.splitext(name)[0])
    ids.sort()
    return ids


def parse_option():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=str, required=True)
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_option()

    with open("../config.yaml", "r") as f:
        cfg = yaml.safe_load(f)

    base_folder = cfg["base_folder"]
    out_folder = cfg["out_folder"]
    ds_cfg = cfg["datasets"][args.dataset]

    meta_dir = os.path.join(out_folder, args.dataset, "our/table_processed/meta_infos")
    xlsx_dir = os.path.join(base_folder, args.dataset, ds_cfg["table_dir"])

    value_index_root = os.path.join(out_folder, args.dataset, "our/value_index")
    ensure_dir(value_index_root)

    # load raw->subtabs mapping if exists
    multitab_mapping_fpath = os.path.join(out_folder, args.dataset, "our/table_processed/multitab_mapping.jsonl")
    raw2subs = load_multitab_mapping_jsonl(multitab_mapping_fpath)

    # discover all available table ids
    all_ids = list_all_table_ids(meta_dir)

    # build raw->subtables groups (with correct "no raw as subtable when children exist")
    raw_to_group = build_raw_to_group(all_ids, raw2subs)

    use_distinct_labels = False if args.dataset == "hitabnum" else True

    raw_ids = sorted(raw_to_group.keys())
    print(f"Discovered {len(raw_ids)} raw tables to process.")

    ok = 0
    failed = 0
    failed_raw_ids = []

    for raw_id in raw_ids:
        sub_ids = sorted(list(dict.fromkeys(raw_to_group[raw_id])))

        # filter to those actually having xlsx+meta
        sub_ids = [
            sid for sid in sub_ids
            if os.path.exists(os.path.join(meta_dir, f"{sid}.json"))
            and os.path.exists(os.path.join(xlsx_dir, f"{sid}.xlsx"))
        ]
        if not sub_ids:
            failed += 1
            failed_raw_ids.append(raw_id)
            continue

        raw_out_dir = os.path.join(value_index_root, raw_id)
        ensure_dir(raw_out_dir)

        # 1) raw-level layered tree (ONE file, nodes+edges)
        write_layered_tree_for_raw(raw_id, sub_ids, meta_dir, raw_out_dir)

        # 2) raw-level value index (ONE value_table + ONE mapping_table)
        build_value_index_for_raw(raw_id, sub_ids, meta_dir, xlsx_dir, raw_out_dir, use_distinct_labels)

        ok += 1
        print(f"[OK] raw={raw_id} (#subtables={len(sub_ids)}) -> {raw_out_dir}")

    print("\n=== Summary ===")
    print(f"OK: {ok}  Failed: {failed}")
    if failed_raw_ids:
        print("Failed raw IDs:", failed_raw_ids)