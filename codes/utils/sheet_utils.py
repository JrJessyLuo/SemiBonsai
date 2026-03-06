import numpy as np
from collections import Counter

import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from bs4 import BeautifulSoup
from copy import copy

from utils.api_utils import *
from utils.constants import *

# 将半结构化的sheet转换为结构化的
def sheet2structure(sheet):
    nrows = sheet.max_row
    ncols = sheet.max_column
    
    for x in range(1, nrows + 1):
        for y in range(1, ncols + 1):
            cell = sheet.cell(row=x, column=y)
            value = cell.value
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            if x1!= x2 or y1!= y2:
                sheet.unmerge_cells(start_row=x1, start_column=y1, end_row=x2, end_column=y2)

            for xx in range(x1, x2 + 1):
                for yy in range(y1, y2 + 1):
                    sheet.cell(row=xx, column=yy, value=value)
    return sheet


# 将Sheet序列化为Markdown格式，必须要求Sheet是结构化的！
def sheet2markdown(sheet):
    # 初始化 Markdown 表格
    markdown_table = ""
    
    # 遍历工作表的每一行
    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        # 如果是第一行，添加表头分隔线
        if index == 0:
            markdown_table += "| " + " | ".join("---" for _ in row) + " |\n"
        else:
            # 将每一行的单元格值转换为 Markdown 格式
            markdown_row = "| " + " | ".join(str(cell) if cell is not None else "" for cell in row) + " |"
            markdown_table += markdown_row + "\n"
        
    return markdown_table

# 将HTMl表格转换为Excel表格
def html2workbook(html_content):
    # 使用 BeautifulSoup 解析 HTML
    soup = BeautifulSoup(html_content, 'html.parser')
    table = soup.find('table')  # 找到表格

    # 创建一个新的 Excel 文件
    wb = openpyxl.Workbook()
    ws = wb.active

    # 首先处理合并单元格的情况
    for row_idx, row in enumerate(table.find_all('tr')):
        for col_idx, cell in enumerate(row.find_all(['td', 'th'])):  # 处理 td 和 th 标签
            # 处理合并单元格
            colspan = 1
            rowspan = 1
            if cell.has_attr('colspan'):  # 列合并
                colspan = int(cell['colspan'])
            if cell.has_attr('rowspan'):  # 行合并
                rowspan = int(cell['rowspan'])
            ws.merge_cells(start_row=row_idx + 1, start_column=col_idx + 1,
                            end_row=row_idx + rowspan, end_column=col_idx + colspan)
    
    # 逐行填入值，如果不是合并单元格的左上角，则往右填值
    x = 1
    for row_idx, row in enumerate(table.find_all('tr')):
        y = 1
        for col_idx, cell in enumerate(row.find_all(['td', 'th'])):  # 处理 td 和 th 标签
            cell_value = cell.get_text(strip=True)
            # 判断当前 x y 位置是否是合并单元格的左上角
            x1, y1, x2, y2 = get_merge_cell_size(ws, ws.cell(row=x, column=y).coordinate)
            while x != x1 and y != y1:
                y += 1
                x1, y1, x2, y2 = get_merge_cell_size(ws, ws.cell(row=x, column=y).coordinate)

            # 填入值
            x1, y1, x2, y2 = get_merge_cell_size(ws, ws.cell(row=x, column=y).coordinate)
            if x1 != x2 or y1 != y2:
                ws.unmerge_cells(start_row=x1, start_column=y1, end_row=x2, end_column=y2)
                ws.cell(row=x, column=y, value=cell_value)
                ws.merge_cells(start_row=x1, start_column=y1, end_row=x2, end_column=y2)
            else:
                ws.cell(row=x, column=y, value=cell_value)
            # 移动到下一列
            y += 1
        x += 1
            
    return wb

def get_xlsx_table_string(table_file):
    sheet = openpyxl.load_workbook(table_file).active
    s = sheet2html(sheet)
    return s

def delete_dict_none_none(data : dict):
    """将JSON中None:None的项去除"""
    new_dict = {}
    for k, v in data.items():
        if (k is None or k == 'None') and (v is None or v == 'None'):
            pass
        else:
            if isinstance(v, dict):
                new_dict[k] = delete_dict_none_none(v)
            elif isinstance(v, list):
                new_v = []
                for item in v:
                    if isinstance(item, dict):
                        new_v.append(delete_dict_none_none(item))
                    else:
                        new_v.append(item)
                new_dict[k] = new_v
            else:
                new_dict[k] = v
    return new_dict

def sheet2html(sheet):
    """Transform a excel sheet into a html file."""
    # 初始化HTML表格
    html = "<table border='1'>\n"

    # 遍历行和列，仅处理在合并单元格左上角的单元格
    nrows = sheet.max_row
    ncols = sheet.max_column

    for row in range(1, nrows + 1):
        html += "  <tr>\n"
        col = 1
        while col <= ncols:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

            col += y2 - y1 + 1
            if x1 != row and y1 != col:
                continue

            rowspan = x2 - x1 + 1
            colspan = y2 - y1 + 1

            # 构建HTML单元格
            html += f"    <td"
            if rowspan > 1:
                html += f" rowspan='{rowspan}'"
            if colspan > 1:
                html += f" colspan='{colspan}'"
            html += f">{cell.value if cell.value else ' '}</td>\n"

        html += "  </tr>\n"

    html += "</table>"

    return html

def delete_empty_rows(sheet):
    """删除空行"""
    rows_to_delete = []
    for row in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        if all(get_merge_cell_value(sheet, cell.coordinate) is None or str(get_merge_cell_value(sheet, cell.coordinate)).strip() == '' for cell in row):
            rows_to_delete.append(row[0].row)

    for row in reversed(rows_to_delete):  # 从下往上删除行，避免影响行号
        sheet.delete_rows(row)


def delete_empty_columns(sheet):
    """删除空列"""
    cols_to_delete = []
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if all(get_merge_cell_value(sheet, cell.coordinate) is None or str(get_merge_cell_value(sheet, cell.coordinate)).strip() == '' for cell in col):
            cols_to_delete.append(col[0].column)

    for col in reversed(cols_to_delete):  # 从右往左删除列，避免影响列号
        sheet.delete_cols(col)


# def get_sub_sheet(sheet, min_row, min_col, max_row, max_col, use_wb=False):
#     """获取工作表的一个子部分，返回一个新的工作簿"""
#     # 不符合坐标规定，返回
#     if min_row > max_row or min_col > max_col:
#         return None
#     # 只有一个单元格且单元格为空，返回
#     if min_row == max_row and min_col == max_col and (sheet.cell(row=min_row, column=min_col) is None or str(sheet.cell(row=min_row, column=min_col)).strip() == ""):
#         return None
    
#     # 定义要提取的子区域的范围（例如，A1到D10）
#     # 创建一个新的工作表
#     wb = openpyxl.Workbook()
#     new_sheet = wb.active

#     # 复制源工作表中的子区域的内容
#     for i, row in enumerate(
#         sheet.iter_rows(
#             min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
#         ),
#         start=1,
#     ):
#         for j, cell in enumerate(row, start=1):
#             # 复制单元格的值
#             new_sheet.cell(row=i, column=j, value=cell.value)

#     # 复制合并单元格
#     for merged_range in sheet.merged_cells.ranges:
#         # 检查合并区域是否完全在指定范围内
#         mmin_col, mmin_row, mmax_col, mmax_row = merged_range.bounds
#         if (
#             mmin_row >= min_row
#             and mmax_row <= max_row
#             and mmin_col >= min_col
#             and mmax_col <= max_col
#         ):

#             new_sheet.merge_cells(
#                 get_coordinate_by_cell_pos(
#                     mmin_row - min_row + 1,
#                     mmin_col - min_col + 1,
#                     mmax_row - min_row + 1,
#                     mmax_col - min_col + 1,
#                 )
#             )

#     if use_wb: return wb
#     return wb.active


def get_sub_sheet(
    sheet,
    min_row,
    min_col,
    max_row,
    max_col,
    *,
    use_wb: bool = False,
    clip_merged: bool = True,   # True: merge 跨边界时裁剪到子区域；False: 只保留完全落在区域内的 merge
    copy_dimensions: bool = True,
):
    """
    尽量“原封不动”复制 sheet 的一个矩形区域到新 Workbook 的新 Sheet。
    - 复制：value/公式、样式、行高列宽、合并单元格
    - clip_merged=True: 合并区域跨边界时裁剪交集（更接近 Excel copy 的结构观感）
    """

    # 参数合法性
    if min_row > max_row or min_col > max_col:
        return None

    # 如果区域内全空（可选：快速判空）
    # 注意：openpyxl 的 cell 对象永远存在，判断空要用 cell.value
    all_empty = True
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            if sheet.cell(row=r, column=c).value not in (None, ""):
                all_empty = False
                break
        if not all_empty:
            break
    if all_empty:
        return None

    wb = openpyxl.Workbook()
    new_sheet = wb.active
    new_sheet.title = f"{sheet.title}_sub"

    # --- 1) 复制单元格内容 + 样式 ---
    for src_r in range(min_row, max_row + 1):
        dst_r = src_r - min_row + 1
        for src_c in range(min_col, max_col + 1):
            dst_c = src_c - min_col + 1

            src_cell = sheet.cell(row=src_r, column=src_c)
            dst_cell = new_sheet.cell(row=dst_r, column=dst_c)

            # 值/公式
            dst_cell.value = src_cell.value

            # 样式（尽量完整）
            # if src_cell.has_style:
            #     dst_cell.font = copy(src_cell.font)
            #     dst_cell.fill = copy(src_cell.fill)
            #     dst_cell.border = copy(src_cell.border)
            #     dst_cell.alignment = copy(src_cell.alignment)
            #     dst_cell.number_format = src_cell.number_format
            #     dst_cell.protection = copy(src_cell.protection)
            #     dst_cell._style = copy(src_cell._style)

            # 其他常见属性
            # if src_cell.hyperlink:
            #     dst_cell.hyperlink = copy(src_cell.hyperlink)
            # if src_cell.comment:
            #     dst_cell.comment = copy(src_cell.comment)

    # --- 2) 复制行高/列宽（非常影响“看起来是否一样”） ---
    if copy_dimensions:
        # 列宽
        for src_c in range(min_col, max_col + 1):
            src_letter = get_column_letter(src_c)
            dst_letter = get_column_letter(src_c - min_col + 1)

            if src_letter in sheet.column_dimensions:
                new_sheet.column_dimensions[dst_letter].width = sheet.column_dimensions[src_letter].width

        # 行高
        for src_r in range(min_row, max_row + 1):
            dst_r = src_r - min_row + 1
            if src_r in sheet.row_dimensions:
                new_sheet.row_dimensions[dst_r].height = sheet.row_dimensions[src_r].height

    # --- 3) 复制合并单元格 ---
    # openpyxl 的 merged range 用 bounds: (min_col, min_row, max_col, max_row)
    for merged_range in sheet.merged_cells.ranges:
        mmin_c, mmin_r, mmax_c, mmax_r = merged_range.bounds

        # 与目标区域是否相交
        inter_min_r = max(mmin_r, min_row)
        inter_max_r = min(mmax_r, max_row)
        inter_min_c = max(mmin_c, min_col)
        inter_max_c = min(mmax_c, max_col)

        if inter_min_r > inter_max_r or inter_min_c > inter_max_c:
            continue  # 不相交

        if not clip_merged:
            # 只保留完全包含在区域内的 merge
            if not (mmin_r >= min_row and mmax_r <= max_row and mmin_c >= min_col and mmax_c <= max_col):
                continue
            use_min_r, use_max_r, use_min_c, use_max_c = mmin_r, mmax_r, mmin_c, mmax_c
        else:
            # 裁剪到子区域交集
            use_min_r, use_max_r, use_min_c, use_max_c = inter_min_r, inter_max_r, inter_min_c, inter_max_c

        # 平移到新 sheet 坐标
        dst_min_r = use_min_r - min_row + 1
        dst_max_r = use_max_r - min_row + 1
        dst_min_c = use_min_c - min_col + 1
        dst_max_c = use_max_c - min_col + 1

        new_sheet.merge_cells(
            start_row=dst_min_r,
            start_column=dst_min_c,
            end_row=dst_max_r,
            end_column=dst_max_c,
        )

    return wb if use_wb else new_sheet


def get_sheet_value_list(sheet):
    """获得一个sheet中所有content的列表，"""
    res = []

    nrows = sheet.max_row
    ncols = sheet.max_column

    for x in range(1, nrows + 1):
        for y in range(1, ncols + 1):
            cell = sheet.cell(row=x, column=y)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            if x != x1 and y != y1:
                continue
            res.append(cell.value)

    return res


def get_merge_cell_size(sheet, cell_pos):
    """给定一个cell, 返回这个cell所在的合并单元格的位置信息[x1, y1, x2, y2]"""
    cell = sheet[cell_pos]
    # 遍历合并单元格范围，检查A11是否在其中
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            # 获取合并单元格的行列跨度
            min_col, min_row, max_col, max_row = merged_range.bounds
            return min_row, min_col, max_row, max_col

    # 不是合并单元格
    return get_cell_pos_by_coordinate(cell_pos)


def get_merge_cell_value(sheet, cell_pos):
    """给定一个cell, 返回这个cell所在的合并单元格的值"""
    x1, y1, x2, y2 = get_merge_cell_size(sheet, cell_pos)
    return sheet.cell(row=x1, column=y1).value


def get_cell_pos_by_coordinate(coordinate_str):
    """根据Cell的Excel坐标表示返回cell的位置
    例如 A21:B32 -> [21, 1, 32, 2] B4 -> [4, 2, 4, 2]
    """

    def excel_column_to_number(column_string):
        """
        将Excel列标（如 'A', 'Z', 'AA', 'AAA'）转换为列号
        """
        column_number = 0
        for char in column_string:
            column_number = column_number * 26 + (ord(char.upper()) - ord("A") + 1)
        return column_number

    if ":" not in coordinate_str:
        start_col_letter = "".join(
            [c for c in coordinate_str if c.isalpha()]
        )  # 提取列字母部分
        start_row = int(
            "".join([c for c in coordinate_str if c.isdigit()])
        )  # 提取行号部分
        start_col = excel_column_to_number(start_col_letter)  # 将列字母转换为列号
        return [start_row, start_col, start_row, start_col]

    start_cell, end_cell = coordinate_str.split(":")

    # 处理起始单元格（例如：AAA123）
    start_col_letter = "".join([c for c in start_cell if c.isalpha()])  # 提取列字母部分
    start_row = int("".join([c for c in start_cell if c.isdigit()]))  # 提取行号部分
    start_col = excel_column_to_number(start_col_letter)  # 将列字母转换为列号

    # 处理结束单元格（例如：AAB123）
    end_col_letter = "".join([c for c in end_cell if c.isalpha()])  # 提取列字母部分
    end_row = int("".join([c for c in end_cell if c.isdigit()]))  # 提取行号部分
    end_col = excel_column_to_number(end_col_letter)  # 将列字母转换为列号

    # 返回坐标格式： [start_row, start_col, end_row, end_col]
    return [start_row, start_col, end_row, end_col]


def get_coordinate_by_cell_pos(x1, y1, x2, y2):
    if x1 == x2 and y1 == y2:
        return f"{get_column_letter(y1)}{x1}"
    # 将列号转换为字母
    col1 = get_column_letter(y1)
    col2 = get_column_letter(y2)

    # 拼接成 Excel 坐标格式
    start_cell = f"{col1}{x1}"
    end_cell = f"{col2}{x2}"

    return f"{start_cell}:{end_cell}"


def single_cell(sheet, x1, y1, x2, y2):
    """判断sheet的子区域是否是一个单元格/完整的合并单元格/合并单元格的一部分还是多个单元格组成"""
    if x1 == x2 and y1 == y2:
        return True

    merged_ranges = sheet.merged_cells.ranges

    for merged_range in merged_ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(merged_range))
        # 判断区域是否完全匹配合并单元格
        if min_r <= x1 and min_c <= y1 and max_r >= x2 and max_c >= y2:
            return True
    return False


def all_merge_row(sheet, x1, y1, x2, y2):
    """判断指定区域里是否所有cell都是merge的整行"""
    if x1 == x2 and y1 == y2:
        return True

    width = y2 - y1 + 1
    for row in range(x1, x2 + 1):
        cell = sheet.cell(row=row, column=y1)
        xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, cell.coordinate)
        if yy2 - yy1 + 1 != width:
            return False
    return True


def all_merge_column(sheet, x1, y1, x2, y2):
    """判断指定区域里是否所有cell都是merge的整列"""
    if x1 == x2 and y1 == y2:
        return True

    height = x2 - x1 + 1
    for col in range(y1, y2 + 1):
        cell = sheet.cell(row=x1, column=col)
        xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, cell.coordinate)
        if xx2 - xx1 + 1 != height:
            return False
    return True


def get_sheet_type_row(schema, sheet):
    """
    4: 一张半结构化表格只会有以下四种情况
    T_LIST: Schema + 属性值列表，只要判断某一列是否存在语义规律，可以使用模型判断，也可以使用规则判断
    T_ATTR: Schema + 属性值，只要判断是否某个属性后仅跟一个单元格
    T_SEMI: Schema + 半结构化表格，只要判断某个属性后的子区域是否不规整
    T_MIX: Schema + 属性值/半结构化表格，只要判断某个属性后仅跟一个单元格
    注意要从最细粒度的schema进行判断
    """
    # 首先判断在Schema后是否存在仅有一个单元格的情况
    nrows = sheet.max_row
    ncols = sheet.max_column

    exist_single_cell = False
    only_single_cell = True
    col = 1
    while col <= ncols:
        cell = sheet.cell(row=1, column=col)
        x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
        width = y2 - y1 + 1
        if all_merge_column(sheet, 1, y1, nrows, y2):
            exist_single_cell = True
        else:
            only_single_cell = False
        col += width

    if only_single_cell:
        return T_ARRT

    # 判断是 T_LIST 还是 T_SEMI
    # 如果一列中所有单元格宽度都等于最细粒度Schema的宽度即是 T_LIST
    type_list = True

    col = 1
    schema_height = schema.max_row
    while col <= ncols:
        schema_cell = schema.cell(row=schema_height, column=col)
        x1, y1, x2, y2 = get_merge_cell_size(schema, schema_cell.coordinate)
        schema_width = y2 - y1 + 1
        row = 1
        while row <= nrows:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            width = y2 - y1 + 1
            height = x2 - x1 + 1
            if width != schema_width:
                type_list = False
                break
            row += height
        col += schema_width

    if type_list:
        return T_LIST

    if not only_single_cell and exist_single_cell:
        return T_MIX

    return T_SEMI


def get_sheet_type_column(schema, sheet):
    """
    4: 一张半结构化表格只会有以下四种情况
    T_LIST: Schema + 属性值列表，只要判断某一列是否存在语义规律，可以使用模型判断，也可以使用规则判断
    T_ATTR: Schema + 属性值，只要判断是否某个属性后仅跟一个单元格
    T_SEMI: Schema + 半结构化表格，只要判断某个属性后的子区域是否不规整
    T_MIX: Schema + 属性值/半结构化表格，只要判断某个属性后仅跟一个单元格
    """
    # 首先判断在Schema后是否存在仅有一个单元格的情况
    nrows = sheet.max_row
    ncols = sheet.max_column

    exist_single_cell = False
    only_single_cell = True
    row = 1
    while row <= nrows:
        cell = sheet.cell(row=row, column=1)
        x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
        height = x2 - x1 + 1
        if all_merge_row(sheet, x1, 1, x2, ncols):
            exist_single_cell = True
        else:
            only_single_cell = False
        row += height

    if only_single_cell:
        return T_ARRT

    # 判断是 T_LIST 还是 T_SEMI
    # 如果一列中所有单元格宽度都等于最细粒度Schema的宽度即是 T_LIST
    type_list = True

    row = 1
    schema_width = schema.max_column
    while row <= nrows:
        schema_cell = schema.cell(row=row, column=schema_width)
        x1, y1, x2, y2 = get_merge_cell_size(schema, schema_cell.coordinate)
        schema_height = x2 - x1 + 1
        col = 1
        while col <= ncols:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)
            width = y2 - y1 + 1
            height = x2 - x1 + 1
            if height != schema_height:
                type_list = False
                break
            col += width
        row += schema_height

    if type_list:
        return T_LIST

    if not only_single_cell and exist_single_cell:
        return T_MIX

    return T_SEMI


def in_pos_list(x, y, pos_list):
    """用于判断某个cell坐标是否在合并单元格范围内"""
    for pos in pos_list:
        if pos[0] <= x and pos[1] <= y and pos[2] >= x and pos[3] >= y:
            return pos
    return None


# 获得Schema的联通区域
def get_adjacent_list(curr_pos, pos_list, traversed, adjacent_list):
    # 遍历 curr_pos 的上下左右
    for x in range(curr_pos[0], curr_pos[2] + 1):  # 遍历左侧
        y = curr_pos[1] - 1
        tar_pos = in_pos_list(x, y, pos_list)
        if tar_pos is not None and traversed[pos_list.index(tar_pos)] is False:
            traversed[pos_list.index(tar_pos)] = True
            adjacent_list.append(tar_pos)
            get_adjacent_list(tar_pos, pos_list, traversed, adjacent_list)
    for x in range(curr_pos[0], curr_pos[2] + 1):  # 遍历右侧
        y = curr_pos[3] + 1
        tar_pos = in_pos_list(x, y, pos_list)
        if tar_pos is not None and traversed[pos_list.index(tar_pos)] is False:
            traversed[pos_list.index(tar_pos)] = True
            adjacent_list.append(tar_pos)
            get_adjacent_list(tar_pos, pos_list, traversed, adjacent_list)
    for y in range(curr_pos[1], curr_pos[3] + 1):  # 遍历上侧
        x = curr_pos[0] - 1
        tar_pos = in_pos_list(x, y, pos_list)
        if tar_pos is not None and traversed[pos_list.index(tar_pos)] is False:
            traversed[pos_list.index(tar_pos)] = True
            adjacent_list.append(tar_pos)
            get_adjacent_list(tar_pos, pos_list, traversed, adjacent_list)
    for y in range(curr_pos[1], curr_pos[3] + 1):  # 遍历下侧
        x = curr_pos[2] + 1
        tar_pos = in_pos_list(x, y, pos_list)
        if tar_pos is not None and traversed[pos_list.index(tar_pos)] is False:
            traversed[pos_list.index(tar_pos)] = True
            adjacent_list.append(tar_pos)
            get_adjacent_list(tar_pos, pos_list, traversed, adjacent_list)
    return adjacent_list, traversed


def get_sheet_type_row_by_pos(sheet, pos_list):
    """
    一张半结构化表格只会有以下四种情况
    T_LIST: Schema + 属性值列表，只要判断某一列是否存在语义规律，可以使用模型判断，也可以使用规则判断
    T_ATTR: Schema + 属性值，只要判断是否某个属性后仅跟一个单元格
    T_SEMI: Schema + 半结构化表格，只要判断某个属性后的子区域是否不规整
    T_OTHER: 不符合以上四种情况的情形 需要使用vlm构建json
    """
    if pos_list is None or len(pos_list) == 0:
        return T_OTHER
    # 首先判断在Schema后是否存在仅有一个单元格的情况
    nrows = sheet.max_row
    ncols = sheet.max_column
    len_schema = len(pos_list)

    traversed = [False] * len_schema
    adjacent_list = []
    for i in range(len(pos_list)):
        if traversed[i]:
            continue
        traversed[i] = True
        res_list, traversed = get_adjacent_list(
            pos_list[i], pos_list, traversed, [pos_list[i]]
        )
        adjacent_list.append(res_list)

    schema_area_list = []
    for adjacent in adjacent_list:
        x1 = 100000
        y1 = 100000
        x2 = 0
        y2 = 0
        for pos in adjacent:
            if x1 >= pos[0] and y1 >= pos[1]:
                x1 = pos[0]
                y1 = pos[1]
            if x2 <= pos[2] and y2 <= pos[3]:
                x2 = pos[2]
                y2 = pos[3]
        schema_area_list.append([x1, y1, x2, y2])

    if len(schema_area_list) == 1:  # Schema 完全联通，说明是LIST或ATTR
        # 判断 LIST，即出现某个schema后面有很多cell，否则是ATTR
        schema_area = schema_area_list[0]
        x = schema_area[2] + 1
        for y in range(schema_area[1], schema_area[3] + 1):
            x1, y1, x2, y2 = get_merge_cell_size(
                sheet, sheet.cell(row=x, column=y).coordinate
            )
            if x2 != nrows:
                return T_LIST
        return T_ARRT
    else:  # 可能是 ATTR 或 SEMI 或 OTHER
        # 判断 ATTR，即多个 schema_area 是平行的
        for i in range(len(schema_area_list) - 1):
            s1 = schema_area_list[i]
            s2 = schema_area_list[i + 1]
            if (s1[0] == s2[0] and s1[2] == s2[2]) or (
                s1[1] == s2[1] and s1[3] == s2[3]
            ):
                for j in range(len(schema_area_list) - 1):
                    schema_area = schema_area_list[j]
                    x = schema_area[2] + 1
                    for y in range(schema_area[1], schema_area[3] + 1):
                        x1, y1, x2, y2 = get_merge_cell_size(
                            sheet, sheet.cell(row=x, column=y).coordinate
                        )
                        if x2 != nrows:
                            return T_LIST
                return T_ARRT
        return T_LIST


def get_sheet_type_column_by_pos(sheet, pos_list):
    """
    一张半结构化表格只会有以下四种情况
    T_LIST: Schema + 属性值列表，只要判断某一列是否存在语义规律，可以使用模型判断，也可以使用规则判断
    T_ATTR: Schema + 属性值，只要判断是否某个属性后仅跟一个单元格
    T_SEMI: Schema + 半结构化表格，只要判断某个属性后的子区域是否不规整
    T_MIX: Schema + 属性值/半结构化表格，只要判断某个属性后仅跟一个单元格
    T_OTHER: 不符合以上四种情况的情形 需要使用vlm构建json
    """
    if pos_list is None or len(pos_list) == 0:
        return T_OTHER
    # 首先判断在Schema后是否存在仅有一个单元格的情况
    nrows = sheet.max_row
    ncols = sheet.max_column
    len_schema = len(pos_list)

    traversed = [False] * len_schema
    adjacent_list = []
    for i in range(len(pos_list)):
        if traversed[i]:
            continue
        traversed[i] = True
        res_list, traversed = get_adjacent_list(
            pos_list[i], pos_list, traversed, [pos_list[i]]
        )
        adjacent_list.append(res_list)

    schema_area_list = []
    for adjacent in adjacent_list:
        x1 = 100000
        y1 = 100000
        x2 = 0
        y2 = 0
        for pos in adjacent:
            if x1 >= pos[0] and y1 >= pos[1]:
                x1 = pos[0]
                y1 = pos[1]
            if x2 <= pos[2] and y2 <= pos[3]:
                x2 = pos[2]
                y2 = pos[3]
        schema_area_list.append([x1, y1, x2, y2])

    if len(schema_area_list) == 1:  # Schema 完全联通，说明是LIST或ATTR
        # 判断 LIST，即出现某个schema后面有很多cell，否则是ATTR
        schema_area = schema_area_list[0]
        y = schema_area[3] + 1
        for x in range(schema_area[0], schema_area[2] + 1):
            x1, y1, x2, y2 = get_merge_cell_size(
                sheet, sheet.cell(row=x, column=y).coordinate
            )
            if y2 != ncols:
                return T_LIST
        return T_ARRT
    else:  # 可能是 ATTR 或 SEMI 或 OTHER
        # 判断 ATTR，即多个 schema_area 是平行的
        for i in range(len(schema_area_list) - 1):
            s1 = schema_area_list[i]
            s2 = schema_area_list[i + 1]
            if (s1[0] == s2[0] and s1[2] == s2[2]) or (
                s1[1] == s2[1] and s1[3] == s2[3]
            ):
                for j in range(len(schema_area_list)):
                    schema_area = schema_area_list[j]
                    y = schema_area[3] + 1
                    for x in range(schema_area[0], schema_area[2] + 1):
                        x1, y1, x2, y2 = get_merge_cell_size(
                            sheet, sheet.cell(row=x, column=y).coordinate
                        )
                        if y2 != ncols:
                            return T_LIST
                return T_ARRT
        return T_SEMI


def build_nested_dict(keys, value):
    if len(keys) == 1:
        return {keys[0]: value}
    return {keys[0]: build_nested_dict(keys[1:], value)}


def merge_json(j1, j2):
    for key, value in j2.items():
        if key in j1 and isinstance(j1[key], dict) and isinstance(value, dict):
            merge_json(j1[key], value)  # 递归合并子字典
        else:
            j1[key] = value  # 直接覆盖或添加新键
    return j1


def print_sheet_type(type_id):
    if type_id == T_LIST:
        print("T_LIST")
    elif type_id == T_ARRT:
        print("T_ATTR")
    elif type_id == T_SEMI:
        print("T_SEMI")
    elif type_id == T_MIX:
        print("T_MIX")


def get_sheet_type(type_id):
    if type_id == T_LIST:
        return "T_LIST"
    elif type_id == T_ARRT:
        return "T_ATTR"
    elif type_id == T_SEMI:
        return "T_SEMI"
    elif type_id == T_MIX:
        return "T_MIX"


def get_ngrams(text, n=3):
    return [text[i : i + n] for i in range(len(text) - n + 1)]


def jaccard_similarity(str1, str2, n=3):
    # 获取 N-gram
    ngrams1 = set(get_ngrams(str1, n))
    ngrams2 = set(get_ngrams(str2, n))

    # 计算交集和并集
    intersection = ngrams1.intersection(ngrams2)
    union = ngrams1.union(ngrams2)

    if len(union) == 0:
        return 0

    # 返回 Jaccard 相似度
    return len(intersection) / len(union)


def is_schema(string, schema_list, n_gram=3, threshod=0.7):
    if string == "":
        return False
    if string in schema_list:
        return True
    for schema in schema_list:
        if schema == "":
            continue
        if jaccard_similarity(string, schema, n_gram) >= threshod:
            return True
    return False


def schema_pos_match(sheet, schema_list, ngram=3, enable_embedding=False):
    nrows = sheet.max_row
    ncols = sheet.max_column

    schema = []
    pos_list = []
    pos2schema = {}

    schema_list = [str(x).strip() for x in schema_list]

    for row in range(1, nrows + 1):
        col = 1
        while col <= ncols:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

            col += y2 - y1 + 1
            if x1 != row and y1 != col:
                continue

            if is_schema(str(cell.value).strip(), schema_list, ngram):
                pos_list.append([x1, y1, x2, y2])
                schema.append(cell.value)
                pos2schema[str([x1, y1, x2, y2])] = cell.value

    return pos_list, schema, pos2schema


def get_schema_direction_by_pos(sheet, pos_list):
    """根据给定的Schema位置信息判断Schema是在上侧还是左侧"""

    # Step 1 获得出现在最顶部Schema的行号和出现在最左侧Schema的列号
    minx = 1000000
    miny = 1000000
    for pos in pos_list:
        minx = min(minx, pos[0])
        miny = min(miny, pos[1])

    # Step 2 分别计算第一行和第一列出现的Schema长度
    first_row_cnt = 0
    first_col_cnt = 0
    for pos in pos_list:
        if pos[0] == minx:
            first_row_cnt += 1
        if pos[1] == miny:
            first_col_cnt += 1

    # Step 3 比较Schema的多少，判断Schema方向
    if first_row_cnt > first_col_cnt:
        return SCHEMA_TOP
    elif first_row_cnt < first_col_cnt:
        return SCHEMA_LEFT
    else:  # first_row_cnt = first_col_cnt
        row_cell = [1000, 1000, 1000, 1000]
        col_cell = [1000, 1000, 1000, 1000]
        for pos in pos_list:
            if pos[0] == minx and (pos[0] <= row_cell[0] or pos[1] <= row_cell[1]):
                row_cell = pos
            if pos[1] == miny and (pos[0] <= col_cell[0] or pos[1] <= col_cell[1]):
                col_cell = pos
        if row_cell[0] < col_cell[0]:
            return SCHEMA_TOP
        elif row_cell[1] > col_cell[1]:
            return SCHEMA_LEFT

    # Step 4 首先根据 (1, 1) 位置的合并单元格是向下合并还是向右合并
    x1, y2, x2, y2 = get_merge_cell_size(sheet, sheet.cell(row=1, column=1).coordinate)
    if x2 > 1:
        return SCHEMA_LEFT
    if y2 > 1:
        return SCHEMA_TOP

    return SCHEMA_LEFT  # 默认按照Schema Left来处理

def granularity_decrease_row(sheet):
    """判断sheet的每一行合并单元格粒度是不是递减的"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    row = 1
    while row <= nrows:
        col = 1
        while col <= ncols:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

            cell2 = sheet.cell(row=row, column=y2 + 1)
            xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, cell2.coordinate)

            if xx2 - xx1 > x2 - x1:
                return False

            col += y2 - y1 + 1
        row += 1
    return True


def granularity_decrease_col(sheet):
    """判断sheet的每一列合并单元格粒度是不是递减的"""
    nrows = sheet.max_row
    ncols = sheet.max_column

    col = 1
    while col <= ncols:
        row = 1
        while row <= nrows:
            cell = sheet.cell(row=row, column=col)
            x1, y1, x2, y2 = get_merge_cell_size(sheet, cell.coordinate)

            cell2 = sheet.cell(row=x2 + 1, column=col)
            xx1, yy1, xx2, yy2 = get_merge_cell_size(sheet, cell2.coordinate)

            if yy2 - yy1 > y2 - y1:
                return False

            row += x2 - x1 + 1
        col += 1
    return True


def transpose_sheet(sheet):

    # 获取工作表的数据
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # 转置数据
    transposed_data = list(map(list, zip(*data)))

    # 创建一个新的工作表来存储转置后的数据
    wb = openpyxl.Workbook()
    transposed_sheet = wb.active

    # 将转置后的数据写入新工作表
    for row in transposed_data:
        transposed_sheet.append(row)

    # 处理合并单元格
    for merge_cell in sheet.merged_cells.ranges:
        # 获取合并单元格的起始和结束行列
        min_row, min_col, max_row, max_col = (
            merge_cell.min_row,
            merge_cell.min_col,
            merge_cell.max_row,
            merge_cell.max_col,
        )
        # 转置合并单元格的行列
        transposed_sheet.merge_cells(
            start_row=min_col, start_column=min_row, end_row=max_col, end_column=max_row
        )

    # 保存工作簿
    # wb.save('./example_transposed.xlsx')

    return transposed_sheet

if __name__ == '__main__':
    import os
    import glob
    import numpy as np
    # from utils import sheet_utils as sheet_u
    from table2tree import extract_excel as ee
    import pandas as pd
    import json
    import pickle

    def hierarchical_depth(obj):
        """
        Depth convention (your spec):
        - Do NOT count the outermost root dict.
        - Each nested container (dict/list) under the root adds +1.
        - A dict whose values are all scalars is a LEAF: contributes +1 and stops.
        - Lists contribute +1 and then recurse into their elements.
        - Scalars add 0.

        Returns the maximum depth under this convention.
        """

        def is_scalar(x):
            return not isinstance(x, (dict, list))

        def depth(x):
            # list: count it (+1) and recurse into elements
            if isinstance(x, list):
                if not x:
                    return 1
                return 1 + max(depth(e) for e in x)

            # dict: if leaf (all scalar values), count it (+1) and stop
            if isinstance(x, dict):
                if not x:
                    return 1
                if all(is_scalar(v) for v in x.values()):
                    return 1
                # non-leaf dict: count it (+1) and go deeper
                return 1 + max(depth(v) for v in x.values())

            # scalar
            return 0

        # Ignore the outermost root dict by not adding +1 for it
        if isinstance(obj, dict):
            if not obj:
                return 0
            return max(depth(v) for v in obj.values())
        else:
            # If the overall JSON isn't a dict, just compute normally
            return depth(obj)


    records = []
    ana_dir = "/research/remote/petabyte/users/s3941200/table_qa_methods/ST-Raptor-new_update/analyse/ana_base"
    table_process_path = "/research/remote/petabyte/users/s3941200/table_qa_methods/ST-Raptor-new_update/cache"
    arrs = []

    for file in glob.glob('/research/remote/petabyte/users/s3941200/table_qa_methods/ST-Raptor-new_update/data/SSTQA-en/table/*.xlsx'):
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = ee.preprocess_sheet(wb.active)
        table_id = file.split('/')[-1].strip('.xlsx')
        # if table_id!='1':continue
        table_json_path = os.path.join(table_process_path, f"{table_id}.json")
        table_json = json.load(open(table_json_path, 'r'))

        table_pkl_path = os.path.join(table_process_path, f"{table_id}.pkl")
        table_pkl = pickle.load(open(table_pkl_path, 'rb'))

        delete_empty_columns(sheet)
        delete_empty_columns(sheet)

        nrows = sheet.max_row
        ncols = sheet.max_column

        nrow_cells = ee.get_nrow_cells(sheet)
        ncol_cells = ee.get_ncol_cells(sheet)
        value_lst = get_sheet_value_list(sheet)

        max_nest_depths = hierarchical_depth(table_json)
        arrs.append(max_nest_depths)


        records.append({
            "table_id": table_id,
            "nrows": nrows,
            "ncols": ncols,
            "nrow_cells": nrow_cells,
            "ncol_cells": ncol_cells,
            "ncnt_values": len(value_lst), 
            "tree_cell_cnt": len(table_pkl.all_value_list()),
            "raw_cell_cnt": len(value_lst),
            "nest_depth": max_nest_depths
        })

    df = pd.DataFrame(records)
    df.to_csv(os.path.join(ana_dir, 'table_basic_size.csv'), index=False)
    # avg = float(pd.DataFrame(records)["ncnt_values"].astype(float).mean())
    # print(avg)
    # avg = float(pd.DataFrame(records)["value_lst"].astype(float).mean())
    # print(avg)
    # print(np.mean(np.array(arrs)))

    # print(records)